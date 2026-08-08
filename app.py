from __future__ import annotations

import io
import hashlib
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo


APP_TITLE = "개인시책 정산 도우미"
LEADER = "허탁"
TARGET_NAMES = [
    "고태준", "공미나", "김재열", "김종수", "박규리", "박병선", "박세아", "박유현",
    "송현주", "신진은", "우장식", "유지원", "윤상필", "이연주", "이준형", "임재웅",
    "전소현", "정민호", "정연태", "정예은", "허민환", "허탁",
]

NAV_NAMES = [LEADER] + sorted(n for n in TARGET_NAMES if n != LEADER)

INSURER_ALIASES = {
    "신한생명": "신한라이프",
    "신한라이프": "신한라이프",
}

NAVY = "172A46"
BLUE = "2F67B1"
LIGHT_BLUE = "EAF1FA"
LIGHT_RED = "FCEBEC"
RED = "B42318"
GREEN = "217346"
GRAY = "667085"
LIGHT_GRAY = "F2F4F7"
WHITE = "FFFFFF"


def D(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def dec_out(value: Any) -> int | float:
    number = D(value)
    return int(number) if number == number.to_integral_value() else float(number)


def money(value: Any) -> str:
    number = D(value)
    if number == number.to_integral_value():
        return f"{int(number):,}원"
    text = format(number, "f").rstrip("0").rstrip(".")
    sign = "-" if text.startswith("-") else ""
    unsigned = text.lstrip("-")
    whole, fraction = (unsigned.split(".") + [""])[:2]
    return f"{sign}{int(whole):,}.{fraction}원"


def normalize_insurer(value: Any) -> str:
    text = str(value or "").strip()
    return INSURER_ALIASES.get(text, text)


def normalize_name(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def summarize_product_name(value: Any, insurer: Any = "") -> str:
    """화면·PDF용 상품명을 핵심 명칭 중심으로 간결하게 정리합니다."""
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return ""
    removable_keywords = (
        "해약환급금", "해지환급금", "무해약환급금", "무해약", "보증비용", "간편가입",
        "간편심사", "일반심사", "간편고지", "일반가입형", "납입면제",
        "세만기", "연만기", "갱신형", "비갱신형", "보증형", "저해지",
    )

    def remove_option_block(match: re.Match) -> str:
        content = match.group(1)
        compact = re.sub(r"\s+", "", content)
        is_revision = bool(re.fullmatch(r"(?:Hi)?\d{2,4}(?:\.\d+)*", compact, flags=re.IGNORECASE))
        is_marker = bool(re.fullmatch(r"[무갱,·/\s]+", content))
        return " " if any(keyword in content for keyword in removable_keywords) or is_revision or is_marker else match.group(0)

    # 약관형 부가 설명은 괄호 모양과 관계없이 제거하고 상품 고유 명칭은 유지합니다.
    for _ in range(2):
        text = re.sub(r"\(([^()]*)\)", remove_option_block, text)
        text = re.sub(r"\[([^\[\]]*)\]", remove_option_block, text)
    text = re.sub(r"^(?:무\)|\(무\))\s*", "", text)
    text = re.sub(r"\(?\s*무배당\s*\)?\d{0,4}", "", text, flags=re.IGNORECASE)
    text = text.split("_")[0]
    text = re.sub(r"-?\s*(?:일반|간편|통합간편)가입형\s*$", "", text)
    text = re.sub(r"(?<=[가-힣A-Za-zⅡⅢIVX])\s*\d{4}$", "", text)
    text = re.sub(r"\(\s*[,·/]?\s*\)|\[\s*[,·/]?\s*\]", "", text)
    text = re.sub(r"\s*[,·/]\s*(?=$)", "", text)
    text = re.sub(r"\s{2,}", " ", text).strip(" _-·,")

    # 보험사 열에 이미 표시되는 회사명은 상품명 앞부분에서만 안전하게 생략합니다.
    prefix_map = {
        "ABL생명": ("ABL생명", "ABL"),
        "DB손보": ("DB손해보험", "DB손보", "DB"),
        "KB손해": ("KB손해보험", "KB손해", "KB"),
        "KB생명": ("KB라이프", "KB생명", "KB"),
        "라이나생명": ("라이나생명", "라이나"),
        "미래에셋생명": ("미래에셋생명", "미래에셋"),
        "삼성생명": ("삼성생명", "삼성"),
        "삼성화재": ("삼성화재", "삼성"),
        "신한라이프": ("신한라이프", "신한생명", "신한"),
        "하나생명": ("하나생명", "하나"),
        "하나손보": ("하나손해보험", "하나손보", "하나"),
        "한화생명": ("한화생명", "한화"),
        "한화손보": ("한화손해보험", "한화손보", "한화"),
        "현대해상": ("현대해상",),
        "흥국화재": ("흥국화재", "흥국", "흥Good"),
    }
    company = normalize_insurer(insurer)
    for prefix in sorted(prefix_map.get(company, ()), key=len, reverse=True):
        if re.match(rf"^{re.escape(prefix)}(?=\s|[가-힣A-Za-z0-9(])", text, flags=re.IGNORECASE):
            text = re.sub(rf"^{re.escape(prefix)}\s*", "", text, count=1, flags=re.IGNORECASE)
            break
    return text


def norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def find_header_row(ws, required: Iterable[str], scan_rows: int = 30) -> tuple[int | None, dict[str, int]]:
    needed = {norm_header(v) for v in required}
    for row in range(1, min(ws.max_row, scan_rows) + 1):
        mapping = {}
        for cell in ws[row]:
            if cell.value is not None:
                mapping[norm_header(cell.value)] = cell.column
        if needed.issubset(mapping):
            return row, mapping
    return None, {}


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    if len(digits) >= 8:
        try:
            return datetime.strptime(digits[:8], "%Y%m%d").date()
        except ValueError:
            return None
    return None


def date_text(value: Any) -> str:
    parsed = parse_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else str(value or "")


def infer_period(texts: Iterable[str]) -> tuple[str, str]:
    payment = ""
    performance = ""
    for text in texts:
        clean = str(text or "")
        patterns = [
            (r"(20\d{2})\s*[.년/-]\s*(\d{1,2})\s*월?\s*(?:지급|지급분)", "payment"),
            (r"(20\d{2})\s*[.년/-]\s*(\d{1,2})\s*월?\s*(?:업적)", "performance"),
        ]
        for pattern, kind in patterns:
            match = re.search(pattern, clean)
            if match:
                formatted = f"{int(match.group(1)):04d}년 {int(match.group(2)):02d}월"
                if kind == "payment" and not payment:
                    payment = formatted
                if kind == "performance" and not performance:
                    performance = formatted
        if not performance:
            match = re.search(r"(20\d{2})[.-](\d{1,2})월?\s*업적", clean)
            if match:
                performance = f"{int(match.group(1)):04d}년 {int(match.group(2)):02d}월"
    return payment, performance


def additional_incentive_month(payment_month: str) -> str:
    """지급월 기준 15차월 추가 시책의 계약 업적월(14개월 전)을 반환합니다."""
    match = re.search(r"(20\d{2})년\s*(\d{1,2})월", str(payment_month or ""))
    if not match:
        return ""
    year, month = int(match.group(1)), int(match.group(2))
    month_index = year * 12 + (month - 1) - 14
    target_year, target_month_index = divmod(month_index, 12)
    return f"{target_year:04d}년 {target_month_index + 1:02d}월"


def policy_key(row: dict[str, Any]) -> tuple[str, str, str]:
    policy = str(row.get("증권번호") or "").strip()
    return (row.get("보험사", ""), policy, row.get("지급/환수 구분", ""))


@dataclass
class AnalysisResult:
    filename: str
    payment_month: str = ""
    performance_month: str = ""
    summary_rows: dict[str, dict[str, Any]] = field(default_factory=dict)
    details: dict[str, list[dict[str, Any]]] = field(default_factory=lambda: {"생보": [], "손보": [], "추가 자체산출": []})
    company_original: dict[str, dict[tuple[str, str], Decimal]] = field(default_factory=lambda: defaultdict(dict))
    warnings: list[dict[str, Any]] = field(default_factory=list)
    missing_sections: list[str] = field(default_factory=list)

    def person_rows(self, name: str, section: str) -> list[dict[str, Any]]:
        return [r for r in self.details.get(section, []) if r.get("모집인") == name]

    def has_activity(self, name: str) -> bool:
        summary = self.summary_rows.get(name, {})
        if D(summary.get("총합계")) != 0 or D(summary.get("실지급액")) != 0:
            return True
        return any(self.person_rows(name, s) for s in self.details)


def classify_detail_sheet(ws, sheet_name: str) -> str | None:
    row, mapping = find_header_row(ws, ["보험사", "모집인", "인정보험료", "지급기준액"])
    if not row:
        return None
    name = sheet_name.replace(" ", "")
    if "추가" in name or "자체산출" in name:
        return "추가 자체산출"
    if "생보" in name or "생명" in name:
        return "생보"
    if "손보" in name or "손해" in name:
        return "손보"
    insurers = [normalize_insurer(ws.cell(r, mapping["보험사"]).value) for r in range(row + 1, min(ws.max_row, row + 30) + 1)]
    if any(any(token in insurer for token in ("손보", "화재", "해상")) for insurer in insurers):
        return "손보"
    return "생보"


def read_detail_sheet(ws, section: str) -> list[dict[str, Any]]:
    header_row, mapping = find_header_row(ws, ["보험사", "모집인", "인정보험료", "지급기준액"])
    if not header_row:
        return []
    wanted = [
        "보험사", "지급/환수구분", "랩포탈ID", "모집인", "원장사용인명", "증권번호", "계약자명",
        "계약일", "시상내용", "인정보험료", "상품명", "비고", "시상률", "지급기준액",
    ]
    aliases = {"지급/환수구분": "지급/환수 구분"}
    records = []
    for row_no in range(header_row + 1, ws.max_row + 1):
        name = normalize_name(ws.cell(row_no, mapping["모집인"]).value)
        if name not in TARGET_NAMES:
            continue
        record: dict[str, Any] = {"구분": section, "원본행": row_no}
        for header in wanted:
            col = mapping.get(header)
            record[aliases.get(header, header)] = ws.cell(row_no, col).value if col else None
        record["모집인"] = name
        record["원장사용인명"] = normalize_name(record.get("원장사용인명"))
        record["보험사 원본"] = str(record.get("보험사") or "").strip()
        record["보험사"] = normalize_insurer(record.get("보험사"))
        record["지급/환수 구분"] = str(record.get("지급/환수 구분") or "").strip()
        record["인정보험료"] = D(record.get("인정보험료"))
        record["시상률"] = D(record.get("시상률"))
        record["지급기준액"] = D(record.get("지급기준액"))
        record["계약일정렬"] = parse_date(record.get("계약일")) or date.max
        records.append(record)
    return records


def read_summary_sheet(ws) -> tuple[dict[str, dict[str, Any]], dict[str, dict[tuple[str, str], Decimal]]]:
    header_row, mapping = find_header_row(ws, ["이름", "총합계", "실지급액"])
    if not header_row:
        return {}, defaultdict(dict)
    name_col = mapping["이름"]
    total_col = mapping["총합계"]
    summaries: dict[str, dict[str, Any]] = {}
    companies: dict[str, dict[tuple[str, str], Decimal]] = defaultdict(dict)
    metadata = {"랩포탈ID", "가이아ID", "소속2", "소속3", "이름", "시상", "재직상태", "해촉일"}
    endings = {"총합계", "생보조정", "손보조정", "기타", "총합계+기타", "소득", "주민세", "실지급액", "변경유무", "비고", "채권압류대상자-"}
    for row_no in range(header_row + 1, ws.max_row + 1):
        name = normalize_name(ws.cell(row_no, name_col).value)
        if name not in TARGET_NAMES:
            continue
        row: dict[str, Any] = {}
        for header, col in mapping.items():
            row[header] = ws.cell(row_no, col).value
        row["이름"] = name
        row["랩포탈ID"] = str(row.get("랩포탈ID") or "")
        for key in ("총합계", "총합계+기타", "소득", "주민세", "실지급액"):
            row[key] = D(row.get(key))
        summaries[name] = row
        for col in range(1, total_col):
            header = str(ws.cell(header_row, col).value or "").strip()
            if not header or header in metadata or header in endings:
                continue
            amount = D(ws.cell(row_no, col).value)
            normalized = normalize_insurer(header)
            kind = "손보" if any(token in normalized for token in ("손보", "손해", "화재", "해상")) else "생보"
            companies[name][(kind, normalized)] = amount
    return summaries, companies


def analyze_workbook(source: bytes | str | Path, filename: str | None = None) -> AnalysisResult:
    stream = io.BytesIO(source) if isinstance(source, bytes) else source
    wb = load_workbook(stream, data_only=True, read_only=False)
    result = AnalysisResult(filename=filename or Path(str(source)).name)
    period_texts = [result.filename]
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
            period_texts.extend(str(v) for v in row if v)
    result.payment_month, result.performance_month = infer_period(period_texts)

    summary_found = False
    found_sections: set[str] = set()
    for ws in wb.worksheets:
        summary_header, _ = find_header_row(ws, ["이름", "총합계", "실지급액"])
        if summary_header and not summary_found:
            result.summary_rows, result.company_original = read_summary_sheet(ws)
            summary_found = True
            continue
        section = classify_detail_sheet(ws, ws.title)
        if section:
            result.details[section].extend(read_detail_sheet(ws, section))
            found_sections.add(section)

    if not summary_found:
        result.missing_sections.append("지급내역_지사")
    for section in ("생보", "손보", "추가 자체산출"):
        if section not in found_sections:
            result.missing_sections.append(section)

    for section in result.details:
        result.details[section].sort(key=lambda r: (r["보험사"], r["계약일정렬"], str(r.get("증권번호") or ""), r["원본행"]))
    build_warnings(result)
    return result


def unique_contracts(rows: list[dict[str, Any]], status: str) -> list[dict[str, Any]]:
    unique: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        if row.get("지급/환수 구분") != status:
            continue
        key = policy_key(row)
        if key not in unique:
            unique[key] = row
    return list(unique.values())


def person_metrics(result: AnalysisResult, name: str) -> dict[str, Any]:
    life = result.person_rows(name, "생보")
    damage = result.person_rows(name, "손보")
    extra = result.person_rows(name, "추가 자체산출")
    regular = life + damage
    paid_contracts = unique_contracts(regular, "지급")
    refund_contracts = unique_contracts(regular, "환수")
    summary = result.summary_rows.get(name, {})
    life_amount = sum((D(r["지급기준액"]) for r in life), Decimal("0"))
    damage_amount = sum((D(r["지급기준액"]) for r in damage), Decimal("0"))
    extra_amount = sum((D(r["지급기준액"]) for r in extra), Decimal("0"))
    detail_total = life_amount + damage_amount + extra_amount
    original_total = D(summary.get("총합계"))
    return {
        "이름": name,
        "랩포탈ID": str(summary.get("랩포탈ID") or next((r.get("랩포탈ID") for r in regular if r.get("랩포탈ID")), "")),
        "지급 건수": len(paid_contracts),
        "지급 인정보험료": sum((D(r["인정보험료"]) for r in paid_contracts), Decimal("0")),
        "환수 건수": len(refund_contracts),
        "환수 인정보험료": sum((D(r["인정보험료"]) for r in refund_contracts), Decimal("0")),
        "생보 시책": life_amount,
        "손보 시책": damage_amount,
        "추가 자체산출": extra_amount,
        "상세 합계": detail_total,
        "총 시책": original_total if summary else detail_total,
        "소득세": D(summary.get("소득")),
        "주민세": D(summary.get("주민세")),
        "실지급액": D(summary.get("실지급액")) if summary else detail_total,
        "차이": original_total - detail_total if summary else Decimal("0"),
        "원본 존재": bool(summary),
        "내역 있음": result.has_activity(name),
    }


def company_rows(result: AnalysisResult, name: str, include_zero_anomalies: bool = False) -> list[dict[str, Any]]:
    detail_regular: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    detail_extra: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    for section in ("생보", "손보"):
        for row in result.person_rows(name, section):
            detail_regular[(section, row["보험사"])] += D(row["지급기준액"])
    for row in result.person_rows(name, "추가 자체산출"):
        detail_extra[("손보", row["보험사"])] += D(row["지급기준액"])
    original = result.company_original.get(name, {})
    keys = set(original) | set(detail_regular) | set(detail_extra)
    output = []
    for key in sorted(keys, key=lambda x: (x[0], x[1])):
        orig = D(original.get(key))
        regular = D(detail_regular.get(key))
        extra = D(detail_extra.get(key))
        final = regular + extra
        diff = orig - final
        if orig == 0 and not (include_zero_anomalies and diff != 0):
            continue
        output.append({
            "구분": key[0], "보험사": key[1], "원본 금액": orig, "일반 상세 산출액": regular,
            "추가 자체산출액": extra, "상세 최종 합계": final, "차이": diff,
        })
    return output


def build_warnings(result: AnalysisResult) -> None:
    warnings: list[dict[str, Any]] = []
    for missing in result.missing_sections:
        if missing == "추가 자체산출":
            continue
        warnings.append({"설계사": "", "구분": "파일 구조", "보험사": "", "내용": f"{missing} 시트를 찾지 못했습니다.", "차이": Decimal("0")})
    for name in NAV_NAMES:
        metrics = person_metrics(result, name)
        if metrics["원본 존재"] and metrics["차이"] != 0:
            warnings.append({"설계사": name, "구분": "개인 총액", "보험사": "", "내용": "원본 총합계와 상세 산출액이 다릅니다.", "차이": metrics["차이"]})
        for company in company_rows(result, name, include_zero_anomalies=True):
            if company["차이"] != 0:
                warnings.append({"설계사": name, "구분": "보험사별 금액", "보험사": company["보험사"], "내용": "보험사별 원본 금액과 상세 산출액이 다릅니다.", "차이": company["차이"]})
    result.warnings = warnings


def summary_dataframe(result: AnalysisResult, include_all: bool = True) -> pd.DataFrame:
    names = NAV_NAMES if include_all else [n for n in NAV_NAMES if result.has_activity(n)]
    rows = []
    warned = {w["설계사"] for w in result.warnings if w["설계사"]}
    for name in names:
        m = person_metrics(result, name)
        row = {k: m[k] for k in ["이름", "지급 건수", "지급 인정보험료", "환수 건수", "환수 인정보험료", "생보 시책", "손보 시책", "추가 자체산출", "총 시책", "소득세", "주민세", "실지급액"]}
        if warned:
            row["확인"] = "확인 필요" if name in warned else ""
        rows.append(row)
    return pd.DataFrame(rows)


DETAIL_COLUMNS = ["모집인", "랩포탈ID", "보험사", "지급/환수 구분", "증권번호", "계약자명", "계약일", "시상내용", "인정보험료", "상품명", "시상률", "지급기준액", "비고"]


def detail_dataframe(result: AnalysisResult, section: str, name: str | None = None, summarize_product: bool = False) -> pd.DataFrame:
    rows = result.details.get(section, [])
    if name:
        rows = [r for r in rows if r["모집인"] == name]
    data = []
    for row in rows:
        data.append({
            col: (
                date_text(row.get(col)) if col == "계약일"
                else dec_out(row.get(col)) if col in ("인정보험료", "시상률", "지급기준액")
                else summarize_product_name(row.get(col), row.get("보험사")) if col == "상품명" and summarize_product
                else row.get(col, "")
            )
            for col in DETAIL_COLUMNS
        })
    return pd.DataFrame(data, columns=DETAIL_COLUMNS)


def export_excel(result: AnalysisResult) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    active_names = [name for name in NAV_NAMES if result.has_activity(name)]
    details = [row for section in ("생보", "손보", "추가 자체산출") for row in result.details.get(section, [])]

    navy_fill = PatternFill("solid", fgColor=NAVY)
    blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    pale_blue_fill = PatternFill("solid", fgColor="F5F8FC")
    red_fill = PatternFill("solid", fgColor=LIGHT_RED)
    amber_fill = PatternFill("solid", fgColor="FFF4CE")
    gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    white_font = Font(name="NanumGothic", size=10, bold=True, color=WHITE)
    body_font = Font(name="NanumGothic", size=10, color=NAVY)
    title_font = Font(name="NanumGothic", size=20, bold=True, color=NAVY)
    section_font = Font(name="NanumGothic", size=12, bold=True, color=NAVY)
    thin = Side(style="thin", color="C9D3E1")
    row_line = Side(style="medium", color="AAB7C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_border = Border(left=thin, right=thin, top=thin, bottom=row_line)
    money_tokens = ("금액", "보험료", "시책", "산출", "합계", "실지급액", "세", "차이")

    def value_out(value: Any) -> Any:
        if isinstance(value, Decimal):
            return dec_out(value)
        return value

    def identifier(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        return re.sub(r"\.0+$", "", text)

    def raw_insurer(row: dict[str, Any]) -> str:
        return str(row.get("보험사 원본") or row.get("보험사") or "").strip()

    def set_money_format(cell, value: Any) -> None:
        number = D(value)
        cell.number_format = '#,##0"원";[Red]-#,##0"원"' if number == number.to_integral_value() else '#,##0.########"원";[Red]-#,##0.########"원"'
        if number < 0:
            cell.font = Font(name="NanumGothic", size=10, color=RED)

    def prepare(ws, freeze: str = "A2") -> None:
        ws.freeze_panes = None
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = ws.page_margins.right = 0.25
        ws.page_margins.top = ws.page_margins.bottom = 0.45

    def style_header(ws, row_no: int, first_col: int, last_col: int) -> None:
        for cell in ws.iter_cols(min_col=first_col, max_col=last_col, min_row=row_no, max_row=row_no):
            c = cell[0]
            c.fill = navy_fill
            c.font = white_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[row_no].height = 30

    def style_body(ws, start: int, end: int, headers: list[str], refund_col: int | None = None) -> None:
        if end < start:
            return
        for row_no in range(start, end + 1):
            refund = refund_col and str(ws.cell(row_no, refund_col).value or "") == "환수"
            for col_no, header in enumerate(headers, 1):
                cell = ws.cell(row_no, col_no)
                cell.font = body_font
                cell.fill = red_fill if refund else (pale_blue_fill if row_no % 2 == 0 else PatternFill("solid", fgColor=WHITE))
                cell.border = row_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if any(token in header for token in money_tokens) and isinstance(cell.value, (int, float)):
                    set_money_format(cell, cell.value)
            ws.row_dimensions[row_no].height = 26

    def add_table(ws, name: str, ref: str) -> None:
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False, showFirstColumn=False, showLastColumn=False)
        ws.add_table(table)

    def internal_link(cell, location: str) -> None:
        """외부 링크 경고 없이 통합 문서 내부 위치로 이동합니다."""
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=location)
        cell.font = Font(name="NanumGothic", size=10, color="0563C1", underline="single")

    def add_person_navigation(ws, names: list[str], title_row: int, block_rows: dict[str, int], columns: int = 7) -> int:
        ws.cell(title_row, 1, "설계사 바로가기").font = section_font
        start_row = title_row + 1
        for index, name in enumerate(names):
            row_no = start_row + index // columns
            col_no = 1 + index % columns
            cell = ws.cell(row_no, col_no, name)
            cell.fill = blue_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # 표 제목보다 아래쪽 셀을 선택해 제목이 화면 상단 쪽에 함께 보이게 합니다.
            internal_link(cell, f"'{ws.title}'!A{block_rows[name] + 6}")
            ws.row_dimensions[row_no].height = 25
        return start_row + ((len(names) - 1) // columns if names else 0)

    def person_block_sheet(title: str, headers: list[str], records_by_name: dict[str, list[list[Any]]], widths: list[float], refund_col: int | None = None, freeze_col: str = "A"):
        ws = wb.create_sheet(title)
        names = [name for name in active_names if records_by_name.get(name)]
        nav_rows = 1 + ((len(names) - 1) // 7 if names else 0)
        first_block = 5 + nav_rows
        block_rows: dict[str, int] = {}
        cursor = first_block
        for name in names:
            block_rows[name] = cursor
            cursor += 2 + len(records_by_name[name]) + 3
        add_person_navigation(ws, names, 1, block_rows)
        table_index = 1
        for name in names:
            title_row = block_rows[name]
            ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(headers))
            title_cell = ws.cell(title_row, 1, f"{name} · {title}")
            title_cell.fill = blue_fill
            title_cell.font = section_font
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            title_cell.border = border
            ws.row_dimensions[title_row].height = 28
            header_at = title_row + 1
            for col, header in enumerate(headers, 1):
                ws.cell(header_at, col, header)
            style_header(ws, header_at, 1, len(headers))
            for values in records_by_name[name]:
                ws.append([]) if ws.max_row < header_at else None
                row_no = ws.max_row + 1
                for col, value in enumerate(values, 1):
                    ws.cell(row_no, col, value_out(value))
            data_end = header_at + len(records_by_name[name])
            style_body(ws, header_at + 1, data_end, headers, refund_col=refund_col)
            add_table(ws, f"T{title.replace('·', '').replace(' ', '')}{table_index}", f"A{header_at}:{get_column_letter(len(headers))}{data_end}")
            table_index += 1
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        prepare(ws, f"{freeze_col}{first_block + 2}")
        return ws

    def contract_groups() -> list[dict[str, Any]]:
        grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
        for row in details:
            policy = identifier(row.get("증권번호"))
            fallback = f"{row.get('구분')}:{row.get('원본행')}" if not policy else policy
            key = (row.get("구분"), row.get("모집인"), identifier(row.get("랩포탈ID")), raw_insurer(row), row.get("보험사"), row.get("지급/환수 구분"), fallback)
            grouped[key].append(row)
        output = []
        for key, rows in grouped.items():
            first = rows[0]
            awards = [str(r.get("시상내용") or "").strip() for r in rows if str(r.get("시상내용") or "").strip()]
            products = [str(r.get("상품명") or "").strip() for r in rows if str(r.get("상품명") or "").strip()]
            notes = [str(r.get("비고") or "").strip() for r in rows if str(r.get("비고") or "").strip()]
            unique_text = lambda values: "\n".join(dict.fromkeys(values))
            output.append({
                "자료구분": key[0], "모집인": key[1], "랩포탈ID": key[2], "보험사": key[3],
                "보험사 정규명": key[4], "지급/환수": key[5], "증권번호": identifier(first.get("증권번호")),
                "계약자명": first.get("계약자명") or "", "계약일": date_text(first.get("계약일")),
                "인정보험료": D(first.get("인정보험료")), "시상 내역 수": len(rows),
                "시상내용": unique_text(awards), "지급기준액 합계": sum((D(r.get("지급기준액")) for r in rows), Decimal("0")),
                "상품명": unique_text(products), "비고": unique_text(notes),
            })
        order = {name: index for index, name in enumerate(NAV_NAMES)}
        return sorted(output, key=lambda r: (order.get(r["모집인"], 999), r["자료구분"], r["보험사"], r["계약일"], r["증권번호"]))

    grouped_contracts = contract_groups()

    # 1. 전체 지급요약
    ws = wb.create_sheet("전체 지급요약")
    ws.merge_cells("A1:N1")
    ws["A1"] = "드림지점 개인시책 지급요약"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    period = " · ".join(v for v in (f"{result.payment_month} 지급" if result.payment_month else "", f"{result.performance_month} 업적" if result.performance_month else "") if v)
    ws.merge_cells("A2:N2")
    ws["A2"] = period or "지급월·업적월 미인식"
    ws["A2"].font = Font(name="NanumGothic", size=11, bold=True, color=GRAY)
    ws["A2"].alignment = Alignment(horizontal="center")
    notice_row = 4
    next_row = 6
    if result.warnings:
        ws.merge_cells(start_row=notice_row, start_column=1, end_row=notice_row, end_column=14)
        ws.cell(notice_row, 1, f"확인 필요한 항목 {len(result.warnings)}건 · 자세한 내용은 ‘확인 필요’ 시트를 확인해 주세요.")
        ws.cell(notice_row, 1).fill = amber_fill
        ws.cell(notice_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        internal_link(ws.cell(notice_row, 1), "'확인 필요'!A1")
        ws.cell(notice_row, 1).font = Font(name="NanumGothic", size=12, bold=True, color="7A4E00", underline="single")
        ws.row_dimensions[notice_row].height = 28
    metrics = [person_metrics(result, name) for name in active_names]
    positive_incentive = sum((m["총 시책"] for m in metrics if D(m["총 시책"]) > 0), Decimal("0"))
    refund_incentive = sum((m["총 시책"] for m in metrics if D(m["총 시책"]) < 0), Decimal("0"))
    cards = [
        ("대상 인원", len(active_names), '0"명"'),
        ("총 계약 건수", sum(m["지급 건수"] + m["환수 건수"] for m in metrics), '0"건"'),
        ("지급 인정보험료", sum((m["지급 인정보험료"] for m in metrics), Decimal("0")), None),
        ("환수 인정보험료", sum((m["환수 인정보험료"] for m in metrics), Decimal("0")), None),
        ("총 시책", positive_incentive, None),
        ("총 환수금액", refund_incentive, None),
        ("실지급액", sum((m["실지급액"] for m in metrics), Decimal("0")), None),
    ]
    for index, (label, value, number_format) in enumerate(cards):
        start = 1 + index * 2
        ws.merge_cells(start_row=next_row, start_column=start, end_row=next_row, end_column=start + 1)
        ws.merge_cells(start_row=next_row + 1, start_column=start, end_row=next_row + 1, end_column=start + 1)
        label_cell, value_cell = ws.cell(next_row, start), ws.cell(next_row + 1, start)
        label_cell.value, value_cell.value = label, value_out(value)
        label_cell.fill = gray_fill
        value_cell.fill = red_fill if label in ("환수 인정보험료", "총 환수금액") else blue_fill
        label_cell.font = Font(name="NanumGothic", size=9, bold=True, color=GRAY)
        value_cell.font = Font(name="NanumGothic", size=13, bold=True, color=NAVY if D(value) >= 0 else RED)
        label_cell.alignment = value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if number_format:
            value_cell.number_format = number_format
        elif isinstance(value, Decimal):
            set_money_format(value_cell, value)
        value_cell.font = Font(name="NanumGothic", size=13, bold=True, color=RED if D(value) < 0 else NAVY)
        for row_no in (next_row, next_row + 1):
            for col_no in (start, start + 1):
                ws.cell(row_no, col_no).border = border
    header_row = next_row + 3
    summary_headers = ["이름", "지급 건수", "지급 인정보험료", "환수 건수", "환수 인정보험료", "생보 시책", "손보 시책", "추가 자체산출", "총 시책", "소득세", "주민세", "실지급액", "확인"]
    ws.append([])
    for col, header in enumerate(summary_headers, 1):
        ws.cell(header_row, col, header)
    style_header(ws, header_row, 1, len(summary_headers))
    warned_names = {w["설계사"] for w in result.warnings if w.get("설계사")}
    for row_no, m in enumerate(metrics, header_row + 1):
        values = [m["이름"], m["지급 건수"], m["지급 인정보험료"], m["환수 건수"], m["환수 인정보험료"], m["생보 시책"], m["손보 시책"], m["추가 자체산출"], m["총 시책"], m["소득세"], m["주민세"], m["실지급액"], "확인 필요" if m["이름"] in warned_names else ""]
        for col, value in enumerate(values, 1):
            ws.cell(row_no, col, value_out(value))
    style_body(ws, header_row + 1, header_row + len(metrics), summary_headers)
    for row_no in range(header_row + 1, header_row + len(metrics) + 1):
        if ws.cell(row_no, 13).value:
            ws.cell(row_no, 13).fill = amber_fill
    total_row = header_row + len(metrics) + 1
    totals = ["합계", sum(m["지급 건수"] for m in metrics), sum((m["지급 인정보험료"] for m in metrics), Decimal("0")), sum(m["환수 건수"] for m in metrics), sum((m["환수 인정보험료"] for m in metrics), Decimal("0")), sum((m["생보 시책"] for m in metrics), Decimal("0")), sum((m["손보 시책"] for m in metrics), Decimal("0")), sum((m["추가 자체산출"] for m in metrics), Decimal("0")), sum((m["총 시책"] for m in metrics), Decimal("0")), sum((m["소득세"] for m in metrics), Decimal("0")), sum((m["주민세"] for m in metrics), Decimal("0")), sum((m["실지급액"] for m in metrics), Decimal("0")), ""]
    for col, value in enumerate(totals, 1):
        cell = ws.cell(total_row, col, value_out(value))
        total_font = Font(name="NanumGothic", size=11.5, bold=True, color=WHITE)
        cell.fill, cell.font, cell.border = navy_fill, total_font, Border(top=Side(style="double", color=NAVY), bottom=Side(style="medium", color=NAVY))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col in (3, 5, 6, 7, 8, 9, 10, 11, 12):
            set_money_format(cell, value)
            cell.font = total_font
    ws.row_dimensions[total_row].height = 28
    add_table(ws, "TOverallSummary", f"A{header_row}:M{header_row + len(metrics)}")
    widths = [12, 11, 17, 11, 17, 15, 15, 17, 15, 13, 13, 15, 12, 12]
    for i, width in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = width
    prepare(ws, f"A{header_row + 1}")

    # 2. 계약별 합산 요약 — 같은 시트 안에서 모집인별 독립 표
    contract_headers = ["자료구분", "모집인", "랩포탈ID", "보험사", "지급/환수", "증권번호", "계약자명", "계약일", "인정보험료", "시상 내역 수", "시상내용", "지급기준액 합계", "상품명"]
    contract_by_name: dict[str, list[list[Any]]] = defaultdict(list)
    for record in grouped_contracts:
        contract_by_name[record["모집인"]].append([value_out(record[h]) for h in contract_headers])
    person_block_sheet("계약별 합산 요약", contract_headers, contract_by_name, [12, 11, 16, 15, 12, 20, 13, 13, 16, 12, 36, 18, 50], refund_col=5, freeze_col="E")

    # 3. 통합 계약상세 — 원본값을 유지하고 모집인별 독립 표
    detail_headers = ["자료구분", "모집인", "랩포탈ID", "보험사", "지급/환수 구분", "증권번호", "계약자명", "계약일", "시상내용", "인정보험료", "시상률", "지급기준액", "상품명", "비고"]
    detail_by_name: dict[str, list[list[Any]]] = defaultdict(list)
    for row in details:
        detail_by_name[row["모집인"]].append([row.get("구분"), row.get("모집인"), identifier(row.get("랩포탈ID")), raw_insurer(row), row.get("지급/환수 구분"), identifier(row.get("증권번호")), row.get("계약자명") or "", row.get("계약일"), row.get("시상내용") or "", dec_out(row.get("인정보험료")), dec_out(row.get("시상률")), dec_out(row.get("지급기준액")), row.get("상품명") or "", row.get("비고") or ""])
    ws = person_block_sheet("통합 계약상세", detail_headers, detail_by_name, [12, 11, 16, 15, 13, 20, 13, 13, 24, 16, 12, 16, 55, 35], refund_col=5, freeze_col="E")
    for row_no in range(1, ws.max_row + 1):
        if ws.cell(row_no, 2).value in active_names:
            ws.cell(row_no, 3).number_format = "@"
            ws.cell(row_no, 6).number_format = "@"
            if isinstance(ws.cell(row_no, 8).value, (date, datetime)):
                ws.cell(row_no, 8).number_format = "yyyy-mm-dd"
            ws.cell(row_no, 11).number_format = "0.########"
            set_money_format(ws.cell(row_no, 12), ws.cell(row_no, 12).value)

    # 4. 환수 관리 — 환수 계약이 있을 때만 생성
    refund_map: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for record in (r for r in grouped_contracts if r["지급/환수"] == "환수"):
        key = (record["모집인"], record["랩포탈ID"], record["보험사 정규명"], record["증권번호"] or f"{record['자료구분']}:{record['계약일']}:{record['계약자명']}")
        if key not in refund_map:
            refund_map[key] = dict(record)
            continue
        current = refund_map[key]
        current["자료구분"] = " · ".join(dict.fromkeys((current["자료구분"] + " · " + record["자료구분"]).split(" · ")))
        current["지급기준액 합계"] += record["지급기준액 합계"]
        for text_key in ("시상내용", "상품명", "비고"):
            current[text_key] = "\n".join(dict.fromkeys(filter(None, (current[text_key] + "\n" + record[text_key]).splitlines())))
    refunds = list(refund_map.values())
    if refunds:
        refund_headers = ["모집인", "랩포탈ID", "보험사", "증권번호", "계약자명", "계약일", "환수 인정보험료", "환수 시책금액", "시상내용", "상품명", "비고"]
        refund_by_name: dict[str, list[list[Any]]] = defaultdict(list)
        for record in sorted(refunds, key=lambda r: (NAV_NAMES.index(r["모집인"]), -abs(r["지급기준액 합계"]))):
            refund_by_name[record["모집인"]].append([record["모집인"], record["랩포탈ID"], record["보험사"], record["증권번호"], record["계약자명"], record["계약일"], dec_out(record["인정보험료"]), dec_out(record["지급기준액 합계"]), record["시상내용"], record["상품명"], record["비고"]])
        ws = person_block_sheet("환수 관리", refund_headers, refund_by_name, [11, 16, 15, 20, 13, 13, 18, 18, 32, 50, 35], freeze_col="D")
        for row_no in range(1, ws.max_row + 1):
            if ws.cell(row_no, 1).value in active_names and ws.cell(row_no, 2).value:
                for cell in ws[row_no][:len(refund_headers)]:
                    cell.fill = red_fill
        for table in ws.tables.values():
            header_at = int(re.search(r"\d+", table.ref.split(":")[0]).group())
            data_end = int(re.search(r"\d+", table.ref.split(":")[1]).group())
            subtotal_row = data_end + 1
            refund_total = sum((D(ws.cell(row_no, 8).value) for row_no in range(header_at + 1, data_end + 1)), Decimal("0"))
            ws.cell(subtotal_row, 7, "소계")
            ws.cell(subtotal_row, 8, dec_out(refund_total))
            set_money_format(ws.cell(subtotal_row, 8), refund_total)
            for cell in ws[subtotal_row][:len(refund_headers)]:
                cell.fill = gray_fill
                cell.font = Font(name="NanumGothic", size=11.5, bold=True, color=NAVY)
                cell.border = row_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(subtotal_row, 8).font = Font(name="NanumGothic", size=11.5, bold=True, color=RED if refund_total < 0 else NAVY)
            ws.row_dimensions[subtotal_row].height = 28

    # 5. 보험사별 최종요약 — 지점 전체 합계 + 설계사별 보험사 합계
    def blank_insurer_metric() -> dict[str, Any]:
        return {
            "지급 계약": set(), "환수 계약": set(), "지급 시책": Decimal("0"),
            "환수 시책": Decimal("0"), "추가 자체산출": Decimal("0"), "설계사": set(),
        }

    insurer_by_name: dict[str, dict[tuple[str, str], dict[str, Any]]] = {
        name: defaultdict(blank_insurer_metric) for name in active_names
    }
    for section in ("생보", "손보"):
        for row in result.details.get(section, []):
            name = row.get("모집인")
            if name not in insurer_by_name:
                continue
            key = (section, row.get("보험사") or raw_insurer(row))
            metric = insurer_by_name[name][key]
            status = row.get("지급/환수 구분")
            policy = identifier(row.get("증권번호")) or f"{section}:{row.get('원본행')}"
            if status == "환수":
                metric["환수 계약"].add(policy)
                metric["환수 시책"] += D(row.get("지급기준액"))
            else:
                metric["지급 계약"].add(policy)
                metric["지급 시책"] += D(row.get("지급기준액"))
            metric["설계사"].add(name)
    for row in result.details.get("추가 자체산출", []):
        name = row.get("모집인")
        if name not in insurer_by_name:
            continue
        key = ("손보", row.get("보험사") or raw_insurer(row))
        insurer_by_name[name][key]["추가 자체산출"] += D(row.get("지급기준액"))
        insurer_by_name[name][key]["설계사"].add(name)

    def metric_row(kind: str, insurer: str, metric: dict[str, Any], include_people: bool) -> dict[str, Any]:
        row = {
            "구분": kind, "보험사": insurer, "지급 계약 건수": len(metric["지급 계약"]),
            "환수 계약 건수": len(metric["환수 계약"]), "지급 시책": metric["지급 시책"],
            "환수 시책": metric["환수 시책"], "추가 자체산출": metric["추가 자체산출"],
            "최종 시책": metric["지급 시책"] + metric["환수 시책"] + metric["추가 자체산출"],
        }
        if include_people:
            row["해당 설계사 수"] = len(metric["설계사"])
        return row

    branch_insurers: dict[tuple[str, str], dict[str, Any]] = defaultdict(blank_insurer_metric)
    for name in active_names:
        for key, metric in insurer_by_name[name].items():
            branch = branch_insurers[key]
            branch["지급 계약"].update((name, policy) for policy in metric["지급 계약"])
            branch["환수 계약"].update((name, policy) for policy in metric["환수 계약"])
            for amount_key in ("지급 시책", "환수 시책", "추가 자체산출"):
                branch[amount_key] += metric[amount_key]
            branch["설계사"].update(metric["설계사"])

    ws = wb.create_sheet("보험사별 최종요약")
    final_headers = ["구분", "보험사", "지급 계약 건수", "환수 계약 건수", "지급 시책", "환수 시책", "추가 자체산출", "최종 시책", "해당 설계사 수"]
    person_headers = final_headers[:-1]
    ws.merge_cells("A1:I1")
    ws["A1"] = "보험사별 최종 시책 요약"
    ws["A1"].font, ws["A1"].alignment = title_font, Alignment(horizontal="center")
    ws.cell(3, 1, "지점 전체 보험사별 합계").font = section_font
    for col, header in enumerate(final_headers, 1):
        ws.cell(4, col, header)
    style_header(ws, 4, 1, len(final_headers))
    branch_rows = [metric_row(kind, insurer, metric, True) for (kind, insurer), metric in sorted(branch_insurers.items())]
    branch_rows = [r for r in branch_rows if any(D(r[k]) != 0 for k in ("지급 시책", "환수 시책", "추가 자체산출", "최종 시책")) or r["지급 계약 건수"] or r["환수 계약 건수"]]
    branch_rows.sort(key=lambda r: (D(r["최종 시책"]), r["보험사"]), reverse=True)
    row_no = 4
    for record in branch_rows:
        row_no += 1
        for col, header in enumerate(final_headers, 1):
            ws.cell(row_no, col, value_out(record[header]))
    style_body(ws, 5, row_no, final_headers)
    if branch_rows:
        add_table(ws, "TInsurerFinalSummary", f"A4:I{row_no}")

    person_insurer_rows: dict[str, list[dict[str, Any]]] = {}
    for name in active_names:
        rows = [metric_row(kind, insurer, metric, False) for (kind, insurer), metric in sorted(insurer_by_name[name].items())]
        person_insurer_rows[name] = [r for r in rows if any(D(r[k]) != 0 for k in ("지급 시책", "환수 시책", "추가 자체산출", "최종 시책")) or r["지급 계약 건수"] or r["환수 계약 건수"]]
        person_insurer_rows[name].sort(key=lambda r: (D(r["최종 시책"]), r["보험사"]), reverse=True)
    nav_title_row = row_no + 3
    nav_line_count = (len(active_names) - 1) // 7 + 1 if active_names else 0
    first_block_row = nav_title_row + nav_line_count + 4
    block_rows: dict[str, int] = {}
    cursor = first_block_row
    for name in active_names:
        block_rows[name] = cursor
        cursor += len(person_insurer_rows[name]) + 6
    add_person_navigation(ws, active_names, nav_title_row, block_rows)
    row_no = first_block_row
    for name in active_names:
        personal_final = sum((D(r["최종 시책"]) for r in person_insurer_rows[name]), Decimal("0"))
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=8)
        ws.cell(row_no, 1, f"{name} · 보험사별 최종 합계 {money(personal_final)} · 랩포탈ID {identifier(person_metrics(result, name)['랩포탈ID'])}")
        ws.cell(row_no, 1).font, ws.cell(row_no, 1).fill = section_font, blue_fill
        ws.cell(row_no, 1).alignment, ws.cell(row_no, 1).border = Alignment(horizontal="left", vertical="center"), border
        row_no += 1
        for col, header in enumerate(person_headers, 1):
            ws.cell(row_no, col, header)
        style_header(ws, row_no, 1, len(person_headers))
        data_start = row_no + 1
        rows = person_insurer_rows[name]
        for record in rows:
            row_no += 1
            for col, header in enumerate(person_headers, 1):
                ws.cell(row_no, col, value_out(record[header]))
        style_body(ws, data_start, row_no, person_headers)
        row_no += 1
        ws.cell(row_no, 1, "소계")
        ws.cell(row_no, 3, sum(r["지급 계약 건수"] for r in rows))
        ws.cell(row_no, 4, sum(r["환수 계약 건수"] for r in rows))
        for col, key in enumerate(("지급 시책", "환수 시책", "추가 자체산출", "최종 시책"), 5):
            value = sum((D(r[key]) for r in rows), Decimal("0"))
            ws.cell(row_no, col, value_out(value))
            set_money_format(ws.cell(row_no, col), value)
        for cell in ws[row_no][:8]:
            cell.fill, cell.font, cell.border, cell.alignment = gray_fill, Font(name="NanumGothic", size=11.5, bold=True, color=NAVY), row_border, Alignment(horizontal="center")
        ws.row_dimensions[row_no].height = 28
        row_no += 4

    for summary_row in range(header_row + 1, header_row + len(metrics) + 1):
        name = wb["전체 지급요약"].cell(summary_row, 1).value
        if name in block_rows:
            name_cell = wb["전체 지급요약"].cell(summary_row, 1)
            internal_link(name_cell, f"'보험사별 최종요약'!A{block_rows[name] + 6}")
            name_cell.border = row_border
            name_cell.fill = pale_blue_fill if summary_row % 2 == 0 else PatternFill("solid", fgColor=WHITE)
            name_cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, width in enumerate([12, 19, 16, 16, 18, 18, 18, 18, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    prepare(ws, "C5")

    # 6. 실제 확인 항목이 있을 때만 생성
    if result.warnings:
        ws = wb.create_sheet("확인 필요")
        ws.merge_cells("A1:E1")
        ws["A1"] = "확인 필요 항목"
        ws["A1"].font, ws["A1"].fill, ws["A1"].alignment = title_font, amber_fill, Alignment(horizontal="center")
        warning_headers = ["설계사", "구분", "보험사", "내용", "차이"]
        for col, header in enumerate(warning_headers, 1): ws.cell(3, col, header)
        style_header(ws, 3, 1, 5)
        for row_no, warning in enumerate(result.warnings, 4):
            for col, header in enumerate(warning_headers, 1): ws.cell(row_no, col, value_out(warning.get(header, "")))
        style_body(ws, 4, ws.max_row, warning_headers)
        for row in ws.iter_rows(min_row=4):
            for cell in row: cell.fill = amber_fill
        add_table(ws, "TWarnings", f"A3:E{ws.max_row}")
        for col, width in enumerate([12, 18, 18, 52, 16], 1): ws.column_dimensions[get_column_letter(col)].width = width
        prepare(ws, "A4")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _pdf_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    font_dir = Path(__file__).resolve().parent / "fonts"
    regular = font_dir / "NanumGothic-Regular.ttf"
    bold = font_dir / "NanumGothic-Bold.ttf"
    if not regular.exists() or not bold.exists():
        raise FileNotFoundError("fonts 폴더에 NanumGothic 글꼴 파일이 필요합니다.")
    try:
        pdfmetrics.getFont("NanumGothic")
    except KeyError:
        pdfmetrics.registerFont(TTFont("NanumGothic", str(regular)))
        pdfmetrics.registerFont(TTFont("NanumGothic-Bold", str(bold)))
        pdfmetrics.registerFontFamily("NanumGothic", normal="NanumGothic", bold="NanumGothic-Bold")
    return "NanumGothic", "NanumGothic-Bold"


def _pdf_styles():
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    gothic, bold = _pdf_fonts()
    styles = getSampleStyleSheet()
    return {
        "brand": ParagraphStyle("brand", parent=styles["Normal"], fontName=bold, fontSize=9.5, textColor=colors.HexColor(f"#{BLUE}"), leading=13),
        "title": ParagraphStyle("title", parent=styles["Title"], fontName=bold, fontSize=20, textColor=colors.HexColor(f"#{NAVY}"), leading=25, spaceAfter=4),
        "sub": ParagraphStyle("sub", parent=styles["Normal"], fontName=bold, fontSize=11, textColor=colors.HexColor("#344054"), leading=16, spaceTop=2, alignment=1),
        "section": ParagraphStyle("section", parent=styles["Heading2"], fontName=bold, fontSize=13, textColor=colors.HexColor(f"#{NAVY}"), leading=17, spaceBefore=11, spaceAfter=6),
        "body": ParagraphStyle("body", parent=styles["Normal"], fontName=gothic, fontSize=7.8, leading=10.7, alignment=1),
        "small": ParagraphStyle("small", parent=styles["Normal"], fontName=gothic, fontSize=7.4, leading=9.5, alignment=1),
        "metric": ParagraphStyle("metric", parent=styles["Normal"], fontName=bold, fontSize=14, leading=18, alignment=1, textColor=colors.HexColor(f"#{NAVY}")),
        "metric_negative": ParagraphStyle("metric_negative", parent=styles["Normal"], fontName=bold, fontSize=14, leading=18, alignment=1, textColor=colors.HexColor("#9B1C1C")),
        "center": ParagraphStyle("center", parent=styles["Normal"], fontName=gothic, fontSize=8, leading=10.5, alignment=1),
        "header": ParagraphStyle("header", parent=styles["Normal"], fontName=bold, fontSize=8, leading=10.5, alignment=1, textColor=colors.white),
        "warning": ParagraphStyle("warning", parent=styles["Normal"], fontName=bold, fontSize=9, leading=12, textColor=colors.HexColor(f"#{RED}")),
    }


def _p(text: Any, style):
    from xml.sax.saxutils import escape
    from reportlab.platypus import Paragraph
    return Paragraph(escape(str(text if text is not None else "")), style)


def _numbered_canvas(label: str):
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    gothic, _ = _pdf_fonts()

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            page_count = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.saveState()
                self.setFont(gothic, 7)
                self.setFillColor(colors.HexColor(f"#{GRAY}"))
                self.drawString(31, 18, label)
                self.drawRightString(self._pagesize[0] - 31, 18, f"{self._pageNumber} / {page_count}")
                self.restoreState()
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)

    return NumberedCanvas


def _summary_pdf(result: AnalysisResult) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle
    styles = _pdf_styles(); buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=12*mm)
    period = " · ".join(v for v in [
        f"{result.payment_month} 지급" if result.payment_month else "",
        f"{result.performance_month} 업적" if result.performance_month else "",
        "13차월 추가 시책",
    ] if v)
    story = [_p("화랑 WORKSPACE", styles["brand"]), _p("전체 지급 요약", styles["title"]), _p(period, styles["sub"]), Spacer(1, 5*mm)]
    df = summary_dataframe(result, include_all=False)
    cols = ["이름", "지급 건수", "지급 인정보험료", "환수 건수", "환수 인정보험료", "생보 시책", "손보 시책", "추가 자체산출", "총 시책", "실지급액"]
    if "확인" in df.columns:
        cols.append("확인")
    data = [[_p(c, styles["header"]) for c in cols]]
    for _, row in df.iterrows():
        vals = []
        for col in cols:
            value = row.get(col, "")
            if col.endswith("건수"):
                value = f"{value}건"
            elif col not in ("이름", "확인"):
                value = money(value)
            vals.append(_p(value, styles["center"]))
        data.append(vals)
    widths = [17, 14, 24, 14, 24, 22, 22, 24, 23, 23] + ([18] if len(cols) == 11 else [])
    table = Table(data, colWidths=[w*mm for w in widths], repeatRows=1)
    commands = [("BACKGROUND", (0,0), (-1,0), colors.HexColor(f"#{NAVY}")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .35, colors.HexColor("#D0D5DD")), ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]), ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5)]
    table.setStyle(TableStyle(commands)); story.append(table)
    doc.build(story, canvasmaker=_numbered_canvas("전체 지급 요약"))
    return buffer.getvalue()


def _person_pdf(result: AnalysisResult, name: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import KeepTogether, PageBreak, SimpleDocTemplate, Spacer, Table, TableStyle
    styles = _pdf_styles(); buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=11*mm, rightMargin=11*mm, topMargin=10*mm, bottomMargin=13*mm)
    m = person_metrics(result, name)
    period = " · ".join(v for v in [
        f"{result.payment_month} 지급" if result.payment_month else "",
        f"{result.performance_month} 업적" if result.performance_month else "",
        "13차월 추가 시책",
    ] if v)
    story = [_p("화랑 WORKSPACE", styles["brand"]), _p(f"{name} 개인시책 정산서", styles["title"]), _p(period, styles["sub"]), Spacer(1, 4*mm)]
    payment_amount = D(m["실지급액"])
    negative_payment = payment_amount < 0
    positive_payment = payment_amount > 0
    if negative_payment:
        payment_style = styles["metric_negative"]
        payment_background = colors.HexColor(f"#{LIGHT_RED}")
    elif positive_payment:
        payment_style = styles["metric"]
        payment_background = colors.HexColor("#DCEAFE")
    else:
        payment_style = styles["center"]
        payment_background = colors.white
    metric_data = [[_p("총 시책", styles["center"]), _p("소득세", styles["center"]), _p("주민세", styles["center"]), _p("실지급액", styles["center"])], [_p(money(m["총 시책"]), styles["metric"]), _p(money(m["소득세"]), styles["center"]), _p(money(m["주민세"]), styles["center"]), _p(money(m['실지급액']), payment_style)]]
    metric = Table(metric_data, colWidths=[45*mm]*4, rowHeights=[8*mm, 14*mm])
    metric_commands = [("BACKGROUND",(0,0),(-1,0),colors.HexColor(f"#{LIGHT_BLUE}")),("BACKGROUND",(3,1),(3,1),payment_background),("GRID",(0,0),(-1,-1),.5,colors.HexColor("#B7C6DA")),("VALIGN",(0,0),(-1,-1),"MIDDLE")]
    if negative_payment:
        metric_commands.append(("BOX",(3,1),(3,1),.7,colors.HexColor("#E3A6A3")))
    metric.setStyle(TableStyle(metric_commands))
    story.append(metric)
    person_warnings = [w for w in result.warnings if w.get("설계사") == name]
    if person_warnings:
        story.append(Spacer(1, 2*mm)); story.append(_p("확인 필요: " + " / ".join(f"{w['보험사']+' ' if w['보험사'] else ''}{w['내용']} 차이 {money(w['차이'])}" for w in person_warnings), styles["warning"]))

    story += [_p("정산 요약", styles["section"])]
    summary_data = [[_p(v, styles["header"]) for v in ["구분","지급 건수","지급 인정보험료","환수 건수","환수 인정보험료","시책 금액"]]]
    for section in ("생보", "손보"):
        rows = result.person_rows(name, section)
        paid = unique_contracts(rows, "지급"); refund = unique_contracts(rows, "환수")
        amount = sum((D(r["지급기준액"]) for r in rows), Decimal("0"))
        summary_data.append([_p(v, styles["center"]) for v in [section, f"{len(paid)}건", money(sum((D(r['인정보험료']) for r in paid), Decimal('0'))), f"{len(refund)}건", money(sum((D(r['인정보험료']) for r in refund), Decimal('0'))), money(amount)]])
    extra = result.person_rows(name, "추가 자체산출")
    summary_data.append([_p(v, styles["center"]) for v in ["추가 자체산출", f"{len({policy_key(r) for r in extra})}건", "중복 제외", "-", "중복 제외", money(m["추가 자체산출"])]])
    table = Table(summary_data, colWidths=[31*mm,22*mm,37*mm,22*mm,37*mm,31*mm], repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor(f"#{NAVY}")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#D0D5DD")),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)])); story.append(table)

    companies = company_rows(result, name)
    if companies:
        story.append(_p("보험사별 원본 지급·환수", styles["section"]))
        has_diff = any(r["차이"] != 0 for r in companies)
        headers = ["구분","보험사","원본 금액"] + (["상세 합계","차이"] if has_diff else [])
        data = [[_p(v, styles["header"]) for v in headers]]
        for row in companies:
            vals = [row["구분"], row["보험사"], money(row["원본 금액"])]
            if has_diff: vals += [money(row["상세 최종 합계"]), money(row["차이"]) if row["차이"] != 0 else ""]
            data.append([_p(v, styles["center"]) for v in vals])
        widths = [28,45,40] + ([40,35] if has_diff else [])
        table = Table(data, colWidths=[w*mm for w in widths], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor(f"#{NAVY}")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#D0D5DD")),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)])); story.append(table)

    for section in ("생보", "손보", "추가 자체산출"):
        rows = result.person_rows(name, section)
        if not rows: continue
        story.append(_p(f"{section} 상세내역", styles["section"]))
        headers = ["보험사","계약자명","계약일","시상내용","인정보험료","지급기준액","상품명"]
        data = [[_p(v, styles["header"]) for v in headers]]
        refund_indexes = []
        for i, row in enumerate(rows, 1):
            vals = [
                row["보험사"], row.get("계약자명", ""), date_text(row.get("계약일")),
                row.get("시상내용", ""), money(row["인정보험료"]),
                money(row["지급기준액"]),
                summarize_product_name(row.get("상품명", ""), row.get("보험사", "")),
            ]
            data.append([_p(v, styles["body"] if j in (3, 6) else styles["small"]) for j,v in enumerate(vals)])
            if row["지급/환수 구분"] == "환수": refund_indexes.append(i)
        widths = [20,18,20,39,23,24,36]
        table = Table(data, colWidths=[w*mm for w in widths], repeatRows=1)
        cmds = [("BACKGROUND",(0,0),(-1,0),colors.HexColor(f"#{NAVY}")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.3,colors.HexColor("#D0D5DD")),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]
        for idx in refund_indexes: cmds.append(("BACKGROUND",(0,idx),(-1,idx),colors.HexColor(f"#{LIGHT_RED}")))
        table.setStyle(TableStyle(cmds)); story.append(table)
    label = f"{name} 개인시책 정산서"
    doc.build(story, canvasmaker=_numbered_canvas(label))
    return buffer.getvalue()


def export_pdf(result: AnalysisResult, name: str | None = None) -> bytes:
    if name:
        return _person_pdf(result, name)
    from pypdf import PdfReader, PdfWriter
    writer = PdfWriter()
    parts = [_summary_pdf(result)] + [_person_pdf(result, n) for n in NAV_NAMES if result.has_activity(n)]
    for part in parts:
        reader = PdfReader(io.BytesIO(part))
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO(); writer.write(out); return out.getvalue()


def export_individual_pdfs_zip(result: AnalysisResult) -> bytes:
    """정산 내역이 있는 설계사의 개인 PDF를 각각 생성해 ZIP으로 묶습니다."""
    buffer = io.BytesIO()
    period_file = re.sub(r"\s+", "", result.payment_month) if result.payment_month else ""
    suffix = f"_{period_file}지급" if period_file else ""
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name in NAV_NAMES:
            if result.has_activity(name):
                archive.writestr(f"{name}_개인시책정산서{suffix}.pdf", _person_pdf(result, name))
    return buffer.getvalue()


def display_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    shown = df.copy()
    for col in shown.columns:
        shown[col] = shown[col].map(
            lambda v: "" if v is None or str(v).strip().lower() in ("nan", "none", "nat") else v
        )
        if col == "증권번호":
            shown[col] = shown[col].map(
                lambda v: re.sub(r"\.0+$", "", str(v).strip()) if v != "" else ""
            )
        if col.endswith("건수"):
            shown[col] = shown[col].map(lambda v: f"{v}건")
        elif col == "시상률":
            shown[col] = shown[col].map(
                lambda v: "" if v is None or str(v).strip() == "" else (format(D(v), "f").rstrip("0").rstrip(".") or "0")
            )
        elif col == "지급기준액" or any(token in col for token in ("보험료", "시책", "실지급액", "세", "금액", "합계", "차이")):
            shown[col] = shown[col].map(money)
    return shown


def centered_styler(df: pd.DataFrame, refund_mask=None):
    """Streamlit 데이터 표의 제목과 본문을 모두 가운데 정렬합니다."""
    styler = (
        df.style
        .set_properties(**{"text-align": "center", "vertical-align": "middle"})
        .set_table_styles([
            {"selector": "th", "props": [("text-align", "center"), ("vertical-align", "middle")]},
        ])
    )
    if refund_mask is not None:
        mask = pd.Series(refund_mask, index=df.index).fillna(False).astype(bool)
        styler = styler.apply(
            lambda row: ["background-color:#FDE8E7;color:#9B1C1C" if mask.loc[row.name] else "" for _ in row],
            axis=1,
        )
    return styler


def render_centered_table(st, df: pd.DataFrame, refund_mask=None, min_width: int = 760, max_height: int | None = None) -> None:
    """모든 셀을 가운데 정렬한 전체 길이 표를 가로 스크롤 형태로 표시합니다."""
    styler = centered_styler(df, refund_mask).format(escape="html").hide(axis="index")
    scroll_class = " is-scrollable" if max_height else ""
    height_style = f";--table-max-height:{max_height}px" if max_height else ""
    st.markdown(
        f'<div class="hw-table-wrap{scroll_class}" style="--table-min-width:{min_width}px{height_style}">{styler.to_html()}</div>',
        unsafe_allow_html=True,
    )


def person_summary_dataframe(result: AnalysisResult, name: str) -> pd.DataFrame:
    rows = []
    for section in ("생보", "손보"):
        details = result.person_rows(name, section)
        paid = unique_contracts(details, "지급")
        refund = unique_contracts(details, "환수")
        rows.append({
            "구분": section,
            "지급 건수": f"{len(paid)}건",
            "지급 인정보험료": money(sum((D(r["인정보험료"]) for r in paid), Decimal("0"))),
            "환수 건수": f"{len(refund)}건",
            "환수 인정보험료": money(sum((D(r["인정보험료"]) for r in refund), Decimal("0"))),
            "시책 금액": money(sum((D(r["지급기준액"]) for r in details), Decimal("0"))),
        })
    extra = result.person_rows(name, "추가 자체산출")
    rows.append({
        "구분": "추가 자체산출",
        "지급 건수": f"{len({policy_key(r) for r in extra})}건",
        "지급 인정보험료": "중복 제외",
        "환수 건수": "-",
        "환수 인정보험료": "중복 제외",
        "시책 금액": money(sum((D(r["지급기준액"]) for r in extra), Decimal("0"))),
    })
    return pd.DataFrame(rows)


def render_app():
    import streamlit as st
    st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")
    st.markdown(f"""
    <style>
    .stApp {{background:#f6f8fb;color:#172a46}}
    header[data-testid="stHeader"],
    [data-testid="stHeader"],
    .stAppHeader {{background:#f6f8fb !important}}
    [data-testid="stToolbar"] {{background:transparent !important}}
    .block-container {{padding-top:1.5rem;max-width:1500px}}
    [data-testid="stSidebar"] {{
        position:relative;
        background:
            radial-gradient(circle at 14% 7%,rgba(83,145,224,.30) 0%,rgba(83,145,224,0) 29%),
            radial-gradient(circle at 92% 88%,rgba(42,91,154,.28) 0%,rgba(42,91,154,0) 36%),
            linear-gradient(155deg,#0d1d32 0%,#172f50 48%,#0f2038 100%);
        border-right:1px solid rgba(157,190,229,.18);
        box-shadow:10px 0 30px rgba(15,31,53,.10);
    }}
    [data-testid="stSidebar"]::before {{
        content:"";position:absolute;inset:0;pointer-events:none;opacity:.20;
        background-image:radial-gradient(rgba(255,255,255,.18) .65px,transparent .65px);
        background-size:15px 15px;
        mask-image:linear-gradient(to bottom,black,transparent 82%);
    }}
    [data-testid="stSidebar"] * {{color:#fff}}
    [data-testid="stSidebarContent"] {{position:relative;z-index:1;padding-top:1.25rem}}
    .side-brand {{
        color:#8fbeef !important;font-size:.70rem;font-weight:800;
        letter-spacing:.11em;margin:0 .15rem .38rem;
    }}
    .side-title {{
        position:relative;color:#fff !important;font-size:1.12rem;font-weight:800;
        padding:.82rem .92rem .82rem 1.05rem;margin-bottom:1.1rem;border-radius:12px;
        border:1px solid rgba(170,204,242,.18);
        background:linear-gradient(135deg,rgba(255,255,255,.11),rgba(255,255,255,.035));
        box-shadow:0 12px 28px rgba(3,13,27,.18),inset 0 1px 0 rgba(255,255,255,.08);
        backdrop-filter:blur(8px);
    }}
    .side-title::before {{
        content:"";position:absolute;left:.52rem;top:.72rem;bottom:.72rem;width:3px;
        border-radius:3px;background:linear-gradient(180deg,#7fb8f4,#2f67b1);
        box-shadow:0 0 12px rgba(104,168,235,.45);
    }}
    .side-section {{
        color:#aebfd5 !important;font-size:.76rem;font-weight:700;
        letter-spacing:.04em;margin:.15rem 0 .45rem;
    }}
    [data-testid="stSidebar"] [role="radiogroup"] {{gap:.16rem}}
    [data-testid="stSidebar"] [role="radiogroup"] label {{
        width:100%;min-height:2rem;padding:.38rem .62rem;border-radius:8px;
        border:1px solid transparent;transition:background .15s ease,border-color .15s ease;
    }}
    [data-testid="stSidebar"] [role="radiogroup"] label > div:first-child {{display:none}}
    [data-testid="stSidebar"] [role="radiogroup"] label:hover {{
        background:linear-gradient(90deg,rgba(255,255,255,.10),rgba(255,255,255,.045));
        border-color:rgba(181,210,242,.13);
    }}
    [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {{
        background:linear-gradient(100deg,#3978c5 0%,#285e9f 100%);
        border-color:rgba(157,201,246,.48);
        box-shadow:0 7px 18px rgba(4,18,36,.25),inset 0 1px 0 rgba(255,255,255,.14);
    }}
    [data-testid="stSidebar"] [role="radiogroup"] label p {{
        margin:0;font-size:.88rem;font-weight:600;
    }}
    .brand {{font-size:.82rem;color:#2f67b1;font-weight:700;letter-spacing:.08em}}
    .title {{font-size:2rem;color:#172a46;font-weight:800;margin:.15rem 0}}
    .sub {{color:#667085;margin-bottom:1.2rem}}
    .metric-card {{background:white;border:1px solid #e4e7ec;border-radius:14px;padding:16px 18px;min-height:92px}}
    .metric-label {{color:#667085;font-size:.82rem}}
    .metric-value {{color:#172a46;font-size:1.35rem;font-weight:800;margin-top:5px}}
    .warning-box {{background:#fff1f1;border-left:4px solid #b42318;padding:12px 14px;border-radius:8px;color:#8d1b13}}
    .download-card-head {{min-height:88px;padding:.12rem .12rem .3rem}}
    .download-card-title {{color:#172a46;font-size:1rem;font-weight:800;margin-bottom:.35rem}}
    .download-card-desc {{color:#667085;font-size:.79rem;line-height:1.45}}
    [data-testid="stVerticalBlockBorderWrapper"] {{
        background:#fff;border-color:#d8dee8 !important;border-radius:14px !important;
        box-shadow:0 4px 14px rgba(23,42,70,.045);
    }}
    .hw-table-wrap {{
        width:100%;overflow-x:auto;margin:.35rem 0 1rem;border:1px solid #d8dee8;
        border-radius:11px;background:#fff;box-shadow:0 2px 8px rgba(23,42,70,.035);
    }}
    .hw-table-wrap.is-scrollable {{max-height:var(--table-max-height);overflow:auto}}
    .hw-table-wrap table {{
        width:100%;min-width:var(--table-min-width);border-collapse:collapse;
        font-size:.86rem;color:#172a46;
    }}
    .hw-table-wrap th,.hw-table-wrap td {{
        text-align:center !important;vertical-align:middle !important;
        padding:.58rem .55rem;border-right:1px solid #d8dee8;border-bottom:1px solid #d8dee8;
        white-space:normal;
    }}
    .hw-table-wrap th {{background:#eaf1fa;color:#415b7a;font-weight:700}}
    .hw-table-wrap thead th {{
        position:sticky;top:0;z-index:3;
        box-shadow:0 1px 0 #cbd5e1,0 3px 7px rgba(23,42,70,.07);
    }}
    .hw-table-wrap tbody tr:last-child td {{border-bottom:1px solid #d8dee8 !important}}
    .hw-table-wrap table {{margin-bottom:0 !important}}
    .hw-table-wrap th:last-child,.hw-table-wrap td:last-child {{border-right:0}}
    [data-testid="stFileUploader"] section,
    [data-testid="stFileUploaderDropzone"] {{
        background:#ffffff !important;
        border:1px dashed #b7c6da !important;
        border-radius:12px !important;
        color:#172a46 !important;
    }}
    [data-testid="stFileUploaderDropzone"] *,
    [data-testid="stFileUploader"] section * {{
        color:#172a46 !important;
    }}
    [data-testid="stFileUploaderDropzone"] button,
    [data-testid="stFileUploader"] section button {{
        background:#f8fafc !important;
        color:#172a46 !important;
        border:1px solid #b7c6da !important;
    }}
    [data-testid="stFileUploaderDropzone"] button:hover,
    [data-testid="stFileUploader"] section button:hover {{
        background:#eaf1fa !important;
        border-color:#2f67b1 !important;
    }}
    </style>""", unsafe_allow_html=True)
    st.markdown('<div class="brand">화랑 WORKSPACE</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="title">{APP_TITLE}</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub">지급 엑셀을 업로드하면 개인시책을 자동 정리하고 정산서를 생성합니다.</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("지급내역 엑셀 업로드", type=["xlsx"], label_visibility="collapsed")
    if not uploaded:
        st.info("XLSX 파일을 업로드해 주세요.")
        return
    uploaded_bytes = uploaded.getvalue()
    file_digest = hashlib.sha256(uploaded_bytes).hexdigest()
    try:
        if st.session_state.get("_analysis_file_digest") != file_digest:
            st.session_state["_analysis_result"] = analyze_workbook(uploaded_bytes, uploaded.name)
            st.session_state["_analysis_file_digest"] = file_digest
            st.session_state["_export_cache"] = {}
        result = st.session_state["_analysis_result"]
    except Exception as exc:
        st.error(f"파일을 분석하지 못했습니다: {exc}")
        return
    if not result.payment_month or not result.performance_month:
        c1, c2 = st.columns(2)
        with c1:
            if not result.payment_month: result.payment_month = st.text_input("지급월 (선택)", placeholder="예: 2026년 06월")
        with c2:
            if not result.performance_month: result.performance_month = st.text_input("업적월 (선택)", placeholder="예: 2026년 05월")

    def cached_output(kind: str, name: str = "") -> bytes:
        cache = st.session_state.setdefault("_export_cache", {})
        cache_key = (file_digest, result.payment_month, result.performance_month, kind, name)
        if cache_key not in cache:
            if kind == "combined_pdf":
                cache[cache_key] = export_pdf(result)
            elif kind == "individual_zip":
                cache[cache_key] = export_individual_pdfs_zip(result)
            elif kind == "excel":
                cache[cache_key] = export_excel(result)
            elif kind == "person_pdf":
                cache[cache_key] = export_pdf(result, name)
            else:
                raise ValueError(f"알 수 없는 출력 형식: {kind}")
        return cache[cache_key]
    option_map = {"전체": "전체", "허탁 · 지점장": "허탁"}
    option_map.update({name: name for name in NAV_NAMES if name != LEADER})
    with st.sidebar:
        st.markdown('<div class="side-brand">화랑 WORKSPACE</div><div class="side-title">개인시책 정산</div><div class="side-section">조회 대상</div>', unsafe_allow_html=True)
        selected_label = st.radio("조회 대상", list(option_map), index=0, label_visibility="collapsed")
        selected = option_map[selected_label]
    global_warnings = [w for w in result.warnings if not w.get("설계사")]
    if global_warnings:
        st.markdown('<div class="warning-box">' + '<br>'.join(w["내용"] for w in global_warnings) + '</div>', unsafe_allow_html=True)

    if selected == "전체":
        df = summary_dataframe(result, include_all=True)
        active = df[df["이름"].map(result.has_activity)]
        cards = [
            ("대상 인원", f"{len(TARGET_NAMES)}명"),
            ("총 계약 건수", f"{int(active['지급 건수'].sum()+active['환수 건수'].sum())}건"),
            ("지급 인정보험료", money(sum((D(v) for v in active["지급 인정보험료"]), Decimal('0')))),
            ("환수 인정보험료", money(sum((D(v) for v in active["환수 인정보험료"]), Decimal('0')))),
            ("총 시책", money(sum((D(v) for v in active["총 시책"]), Decimal('0')))),
            ("실지급액", money(sum((D(v) for v in active["실지급액"]), Decimal('0')))),
        ]
        columns = st.columns(6)
        for col, (label, value) in zip(columns, cards):
            col.markdown(f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{value}</div></div>', unsafe_allow_html=True)
        st.markdown("### 전체 요약")
        sort = st.selectbox("정렬", ["이름 가나다순", "총 시책 높은 순", "총 시책 낮은 순"], label_visibility="collapsed")
        if sort == "이름 가나다순":
            df = df.sort_values("이름", ascending=True, kind="stable")
        else:
            mapping = {"총 시책 높은 순":("총 시책",False),"총 시책 낮은 순":("총 시책",True)}
            col, asc = mapping[sort]; df = df.sort_values(col, ascending=asc, kind="stable")
        render_centered_table(st, display_dataframe(df), min_width=1220, max_height=680)
        period_file = re.sub(r"\s+", "", result.payment_month) if result.payment_month else ""
        st.markdown("### 결과 다운로드")
        c1, c2, c3 = st.columns(3)
        with c1:
            with st.container(border=True):
                st.markdown('<div class="download-card-head"><div class="download-card-title">📘 통합 정산서 PDF</div><div class="download-card-desc">전체 지급 요약과 모든 개인 정산서를<br>하나의 PDF로 내려받습니다.</div></div>', unsafe_allow_html=True)
                st.download_button("📥 통합 PDF 다운로드", cached_output("combined_pdf"), file_name=f"드림지점_개인시책정산서_{period_file+'지급' if period_file else ''}.pdf", mime="application/pdf", use_container_width=True)
        with c2:
            with st.container(border=True):
                st.markdown('<div class="download-card-head"><div class="download-card-title">📦 설계사별 PDF ZIP</div><div class="download-card-desc">내역이 있는 설계사의 정산서를<br>개별 PDF로 묶어 내려받습니다.</div></div>', unsafe_allow_html=True)
                st.download_button("📥 개별 PDF ZIP 다운로드", cached_output("individual_zip"), file_name=f"드림지점_설계사별_개인시책정산서_{period_file+'지급' if period_file else ''}.zip", mime="application/zip", use_container_width=True)
        with c3:
            with st.container(border=True):
                st.markdown('<div class="download-card-head"><div class="download-card-title">📊 지급 요약 Excel</div><div class="download-card-desc">전체 요약과 계약 상세내역,<br>확인이 필요한 검증 결과를 포함합니다.</div></div>', unsafe_allow_html=True)
                st.download_button("📥 요약 Excel 다운로드", cached_output("excel"), file_name=f"드림지점_개인시책지급요약_{period_file+'지급' if period_file else ''}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    else:
        st.markdown(f"### {selected} 개인 정산")
        if not result.has_activity(selected):
            st.info(f"{selected} 설계사의 요약할 정산 내역이 없습니다.")
            return
        m = person_metrics(result, selected)
        cards = [("총 시책", money(m["총 시책"])),("소득세",money(m["소득세"])),("주민세",money(m["주민세"])),("실지급액",money(m["실지급액"]))]
        cols=st.columns(4)
        for col,(label,value) in zip(cols,cards): col.markdown(f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{value}</div></div>',unsafe_allow_html=True)
        personal_warnings=[w for w in result.warnings if w.get("설계사")==selected]
        if personal_warnings: st.markdown('<div class="warning-box">'+'<br>'.join(f"{w['보험사']+' · ' if w['보험사'] else ''}{w['내용']} 차이 {money(w['차이'])}" for w in personal_warnings)+'</div>',unsafe_allow_html=True)
        st.markdown("#### 정산 요약")
        render_centered_table(st, person_summary_dataframe(result, selected), min_width=760)
        companies=company_rows(result,selected)
        if companies:
            st.markdown("#### 보험사별 원본 지급·환수")
            cdf=pd.DataFrame(companies)
            if not any(cdf["차이"].map(D)!=0): cdf=cdf[["구분","보험사","원본 금액"]]
            render_centered_table(st, display_dataframe(cdf), min_width=620)
        for section in ("생보","손보","추가 자체산출"):
            st.markdown(f"#### {section} 상세내역")
            ddf=detail_dataframe(result,section,selected,summarize_product=False)
            if ddf.empty: st.caption(f"{section} 내역이 없습니다.")
            else:
                refund_mask = ddf["지급/환수 구분"].eq("환수")
                screen_columns = ["보험사","지급/환수 구분","증권번호","계약자명","계약일","시상내용","인정보험료","시상률","지급기준액","상품명","비고"]
                shown = display_dataframe(ddf[screen_columns]).rename(columns={"지급/환수 구분":"지급/환수"})
                render_centered_table(st, shown, refund_mask, min_width=1380, max_height=680)
        period_file=re.sub(r"\s+","",result.payment_month) if result.payment_month else ""
        st.download_button(f"{selected} 개인정산서 PDF",cached_output("person_pdf", selected),file_name=f"{selected}_개인시책정산서_{period_file+'지급' if period_file else ''}.pdf",mime="application/pdf",use_container_width=True)


if __name__ == "__main__":
    render_app()
